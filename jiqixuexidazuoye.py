from openpyxl import load_workbook
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import xlwt as xlwt
from sklearn.feature_selection import mutual_info_regression
from lightgbm import LGBMClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
from sklearn.model_selection import RepeatedStratifiedKFold, cross_val_score

# 导入训练和测试数据
data = pd.read_csv('train.csv')
data_test = pd.read_excel('test.xlsx')
print(data.info())
print(data_test.info())

# #查看婴儿健康情况
# fig = plt.figure()
# data_train['fetal_health'].value_counts().plot(kind='bar')
# plt.ylabel("number")
# plt.xlabel("fetal_health")
# plt.title("yingerjiankangfenbu")
# plt.show()

'''# #查看各个因素分布
fig = plt.figure()
plt.subplot2grid((7,3),(0,0))
data_train['baseline value'].value_counts().plot(kind='bar')
plt.ylabel("count")
plt.xlabel("baseline value")

plt.subplot2grid((7,3),(0,1))
data_train['accelerations'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu2")

plt.subplot2grid((7,3),(0,2))
data_train['fetal_movement'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu3")

plt.subplot2grid((7,3),(1,0))
data_train['uterine_contractions'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu4")

plt.subplot2grid((7,3),(1,1))
data_train['light_decelerations'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu5")

plt.subplot2grid((7,3),(1,2))
data_train['severe_decelerations'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu6")

plt.subplot2grid((7,3),(2,0))
data_train['prolongued_decelerations'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu7")

plt.subplot2grid((7,3),(2,1))
data_train['abnormal_short_term_variability'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu8")

plt.subplot2grid((7,3),(2,2))
data_train['mean_value_of_short_term_variability'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu9")

plt.subplot2grid((7,3),(3,0))
data_train['percentage_of_time_with_abnormal_long_term_variability'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu10")

plt.subplot2grid((7,3),(3,1))
data_train['mean_value_of_long_term_variability'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu11")

plt.subplot2grid((7,3),(3,2))
data_train['histogram_width'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu12")

plt.subplot2grid((7,3),(4,0))
data_train['histogram_min'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu13")

plt.subplot2grid((7,3),(4,1))
data_train['histogram_max'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu14")

plt.subplot2grid((7,3),(4,2))
data_train['histogram_number_of_peaks'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu15")

plt.subplot2grid((7,3),(5,0))
data_train['histogram_number_of_zeroes'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu16")

plt.subplot2grid((7,3),(5,1))
data_train['histogram_mode'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu17")

plt.subplot2grid((7,3),(5,2))
data_train['histogram_mean'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu18")

plt.subplot2grid((7,3),(6,0))
data_train['histogram_median'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu19")

plt.subplot2grid((7,3),(6,1))
data_train['histogram_variance'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu20")

plt.subplot2grid((7,3),(6,2))
data_train['histogram_tendency'].value_counts().plot(kind='bar')
plt.ylabel("number")
plt.xlabel("yinsu21")
plt.show()'''

data = data.drop('number', axis=1)
'''mi = pd.DataFrame(mutual_info_regression(data[data.columns.drop('fetal_health')],
                                         data[['fetal_health']]),
                  columns=['MI Scores'],
                  index=data.columns.drop('fetal_health'))
corr = pd.DataFrame(data[data.columns.drop('fetal_health')].corrwith(data['fetal_health']),
                    columns=['Correlation'])
s_corr = pd.DataFrame(data[data.columns.drop('fetal_health')].corrwith(data['fetal_health'],
                                                                       method='spearman'),
                      columns=['Spearman_Correlation'])
relation = mi.join(corr)
relation = relation.join(s_corr)
relation=relation.sort_values(by='MI Scores', ascending=False)
print(relation)'''
# #查看前五个特征分布
'''plt.figure(figsize=(12, 5))
sns.boxenplot(data=data, x='fetal_health', y='histogram_variance')
plt.title('histogram_variance by Fetal Health Groups', fontsize=20)
plt.xlabel('Fetal Health', fontsize=15)
plt.ylabel('histogram_variance', fontsize=15)
plt.show()'''
f_train = data[['percentage_of_time_with_abnormal_long_term_variability', 'histogram_min',
    'abnormal_short_term_variability', 'baseline value', 'histogram_variance']]
re_train = data_test[['percentage_of_time_with_abnormal_long_term_variability', 'histogram_min',
    'abnormal_short_term_variability', 'baseline value', 'histogram_variance']]
f_true = data[['fetal_health']]
model = LGBMClassifier()
model.fit(f_train, f_true)
re_pred = model.predict(re_train)
workbook = xlwt.Workbook(encoding='utf-8')
bg = load_workbook('predict.xlsx')
sheet = bg["Sheet1"]
for i in range(len(re_pred)):
    sheet.cell(i + 2, 4, re_pred[i])
bg.save("predict.xlsx")
# #训练及预测 X_为训练特征，Y-为预测结果
'''X_train = data.drop('fetal_health', axis=1)
y_train = data[['fetal_health']]
X_valid = data_test.drop('number', axis=1)
model = LGBMClassifier()
model.fit(X_train, y_train)
y_test_pred = model.predict(X_train)
y_pred = model.predict(X_valid)
print(y_pred)'''
# 写入文件
# workbook = xlwt.Workbook(encoding='utf-8')
# bg = load_workbook('predict.xlsx')
# sheet = bg["Sheet1"]
# for i in range(len(y_pred)):
#     sheet.cell(i + 2, 2, y_pred[i])
# bg.save("predict.xlsx")
